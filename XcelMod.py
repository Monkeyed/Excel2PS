# Program for modifying excel tables of categories and products for Prestashop import

import os

# 0.1 imp necessary modules
import openpyxl

# import time 

print('imported openpyxl, os')
# os.chdir('C:\Users\konra\Documents\GitHub\Ukens Dental Nordent to PS transfer\XcelMod.py')

# newTable = openpyxl.load_workbook(input()) [only when multiple tables in folder]
wb = openpyxl.load_workbook('files\Cats.xlsx')
print('table has been imported to wb')

print('current table has following sheets:')
print(wb.sheetnames)
sheet = wb['Sheet1']

# Categories Functions


def catCorrMetaTitle():
    # Rename all cell elements in Column 'H' to have ' von Nordent' behind them+++++++
    for x in range(1, len(sheet['H'])):
        oldTitle = str(sheet['H'][x].value)
        newTitle = str(oldTitle) + ' von Nordent'
        sheet['H'][x].value = newTitle
    print('1.) Column H: Meta Titles adjusted.')


def catCorrCatParents():
    # Rename parent_id according to categories_name column 'G' of respective parent++++++
    for x in range(1, len(sheet['C'])):
        oldParent_id = int(sheet['C'][x].value)
        if oldParent_id == 0:
            sheet['C'][x].value = str('Home')
            # print(str(sheet['C'][x].value))
        else:
            sheet['C'][x].value = str(sheet['G'][oldParent_id - 1].value)
            # print(str(sheet['C'][x].value))
    print('2.) Column C: Category Parents changed from ID# to actual parent names.')


def catCorrFriendlyURLs():
    # Rename categories_meta_keywords to friendly URL title ++++++++
    for x in range(1, len(sheet['J'])):
        if "," in str(sheet['J'][x].value):
            sheet['J'][x].value = str(sheet['J'][x].value).replace(",", "")
        if " / " in str(sheet['J'][x].value):
            sheet['J'][x].value = str(sheet['J'][x].value).replace(" / ", "")
        if " - " in str(sheet['J'][x].value):
            sheet['J'][x].value = str(sheet['J'][x].value).replace(" - ", " ")
        sheet['J'][x].value = str('Nordent ' + sheet['J'][x].value)
        # print(str(sheet['J'][x].value))
    print('3.) Column J: can now be used as friendly URL links.')


def catCorrImgLinks():
    # Rename img links to have rel. path (/html/ukens-dental/img/nordent_de_cat_images/+)
    for x in range(1, len(sheet['F'])):
        oldImg = str(sheet['F'][x].value)
        sheet['F'][x].value = str('https://ukens-dental.de/img/nordent_de_cat_images/' + oldImg)
        # print(str(sheet['F'][x].value) + '.....' + str(sheet['A'][x].value))
    print('4.) Column F: img links point to: https://ukens-dental.de/img/nordent_de_cat_images/ + xxx.jpg.')


def catCorrCatIds():
    # Rewrite Categories to be 1xxx, so Nordent is in thousands, whereas Calset is then 2000
    for x in range(1, len(sheet['A'])):
        sheet['A'][x].value = int(sheet['A'][x].value) + 1000
        # print(str(sheet['A'][x].value))
    print('5.) Column A: Category IDs have been +1000, now range from 1002 to 1126.')


def catCorrMetaDesc():
    # Make Meta Descriptions same as Category Titles from H, and add description text
    for x in range(1, len(sheet['I'])):
        sheet['I'][x].value = str(str(
            sheet['H'][x].value) + ' - Langlebige Dentalinstrumente und Zubehör, exklusiv erhältlich bei Ukens Dental')
    print('6.) Column I: Adjusted Meta Descriptions. ')


def catHeaderImgToCatImg():
    # Copy column F to B for absolute image paths
    for x in range(1, len(sheet['A'])):
        sheet['B'][x].value = str(sheet['F'][x].value)
        # print(str(sheet['B'][x].value))
        # print('7.) Column B: now same abs Paths as F')


def catSaveFile():
    # save of all changes to new file
    wb.save('Cats_edited.xlsx')


def runCatFunction():
    catCorrMetaTitle()
    catCorrCatParents()
    catCorrFriendlyURLs()
    catCorrImgLinks()
    catCorrCatIds()
    catCorrMetaDesc()
    catHeaderImgToCatImg()

    catSaveFile()


runCatFunction()
