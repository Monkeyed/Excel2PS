# Program for modifying excel tables of categories and products for Prestashop import

import os
import openpyxl
import re


# Beginning Categories Functions

cats_wb = openpyxl.load_workbook('files\cats.xlsx')
cats = cats_wb['Sheet1']

# Rename all cell elements in Column 'H' to have ' von Nordent' behind them
def cat_meta_title():
    
    for x in range(1, len(cats['H'])):
        oldTitle = str(cats['H'][x].value)
        newTitle = str(oldTitle) + ' von Nordent'
        cats['H'][x].value = newTitle
    #print('1.) Column H: Meta Titles adjusted.')

# Rename parent_id according to categories_name column 'G' of respective parent
def cat_corr_parents():
    
    for x in range(1, len(cats['C'])):
        oldParent_id = int(cats['C'][x].value)
        if oldParent_id == 0:
            cats['C'][x].value = str('Home')
            # print(str(cats['C'][x].value))
        else:
            cats['C'][x].value = str(cats['G'][oldParent_id - 1].value)
            # print(str(cats['C'][x].value))
    #print('2.) Column C: Category Parents changed from ID# to actual parent names.')

# Rename categories_meta_keywords to friendly URL title 
def cat_friendly_urls():
    
    for x in range(1, len(cats['J'])):
        if "," in str(cats['J'][x].value):
            cats['J'][x].value = str(cats['J'][x].value).replace(",", "")
        if " / " in str(cats['J'][x].value):
            cats['J'][x].value = str(cats['J'][x].value).replace(" / ", "")
        if " - " in str(cats['J'][x].value):
            cats['J'][x].value = str(cats['J'][x].value).replace(" - ", " ")
        cats['J'][x].value = str('Nordent ' + cats['J'][x].value)
        # print(str(cats['J'][x].value))
    #print('3.) Column J: can now be used as friendly URL links.')

# Rename img links to have rel. path (/html/ukens-dental/img/nordent_de_cat_images/+)
def cat_img_links():
    
    for x in range(1, len(cats['F'])):
        oldImg = str(cats['F'][x].value)
        cats['F'][x].value = str('https://ukens-dental.de/img/nordent_de_cat_images/' + oldImg)
        # print(str(cats['F'][x].value) + '.....' + str(cats['A'][x].value))
    #print('4.) Column F: img links point to: https://ukens-dental.de/img/nordent_de_cat_images/ + xxx.jpg.')

# Rewrite Categories to be 1xxx, so Nordent is in thousands, whereas Calset is then 2000 (future)
def cat_corr_cat_ids():
    
    for x in range(1, len(cats['A'])):
        cats['A'][x].value = int(cats['A'][x].value) + 1000
        # print(str(cats['A'][x].value))
    #print('5.) Column A: Category IDs have been +1000, now range from 1002 to 1126.')


# Make Meta Descriptions same as Category Titles from H, and add description text'
def cat_meta_desc():    
    for x in range(1, len(cats['I'])):
        cats['I'][x].value = str(str(
            cats['H'][x].value) + ' - Langlebige Dentalinstrumente und Zubehör, exklusiv erhältlich bei Ukens Dental')
    #print('6.) Column I: Adjusted Meta Descriptions. ')

# Copy column F to B for absolute image paths
def cat_headimg_to_catimg():
    
    for x in range(1, len(cats['A'])):
        cats['B'][x].value = str(cats['F'][x].value)
        # print(str(cats['B'][x].value))
        # print('7.) Column B: now same abs Paths as F')


def run_cat_funs():
    cat_meta_title()
    cat_corr_parents()
    cat_friendly_urls()
    cat_img_links()
    cat_corr_cat_ids()
    cat_meta_desc()
    cat_headimg_to_catimg()
    cats_wb.save('files\cats_edited.xlsx')

print('running category functions on cats_edited.xlsx ...')
#run_cat_funs()
print('category functions done.')


# End Category Functions

#---------------------------------------------------------------------------
#---------------------------------------------------------------------------

# Beginning image functions

print('')

print('running image functions on products_images.xlsx ...')  

wb_imgs = openpyxl.load_workbook('files\products_images.xlsx')
imgs = wb_imgs['Sheet1']

# import listing of all images into second list of object 'files', from os.walk() function

for files in os.walk("files\product_images"):
    img_list = files[2]

#image list now copied to list img_list

#extracting product id from image name and copying to column C, product_id
def imgs_to_prods():
    for x in range(0, len(img_list)):

        # copying image names to first column after column name
        imgs['A'][x+1].value = img_list[x]

        # splitting image names by underscore to extract product id
        name_string = img_list[x]
        img_name = name_string.partition("_")
        # print(img_name)

        # copying image names to second column after column name
        imgs['C'][x+1].value = str(img_name[0])


#function to make lists of image names and urls
def all_lists_and_names():
    img_list_len = len(img_list)
    for x in range(0, img_list_len):
        curr_id = imgs['C'][x+1].value
        curr_img_name_list = []
        curr_img_name_list.append(str(imgs['A'][x+1].value))
        imgs['D'][x+1].value = None
        #adding url function to make it only run once
        curr_img_url_list = []
        curr_img_url_list.append(str(imgs['B'][x+1].value))
        imgs['E'][x+1].value = None
        print(curr_id)
        for y in range(0, img_list_len):
            if curr_id == imgs['C'][y+1].value and imgs['A'][y+1].value not in curr_img_name_list:
                curr_img_name_list.append(str(imgs['A'][y+1].value))
            #from url function
            if curr_id == imgs['C'][y+1].value and imgs['B'][y+1].value not in curr_img_url_list:
                curr_img_url_list.append(str(imgs['B'][y+1].value))
        for d in range (0, len(curr_img_name_list)):
            if d == 0:
                imgs['D'][x+1].value = str(curr_img_name_list[d])
            else:
                imgs['D'][x+1].value = str(str(imgs['D'][x+1].value) + ',' + str(curr_img_name_list[d]))
        #from url function
        for e in range (0, len(curr_img_url_list)):
            if e == 0:
                imgs['E'][x+1].value = str(curr_img_url_list[e])
            else:
                imgs['E'][x+1].value = str(str(imgs['E'][x+1].value) + ',' + str(curr_img_url_list[e]))
        print('image name and url list done for: ' + str(imgs['C'][x+1].value))

def all_imgs_urls_to_single_product_id():
    for x in range(0, len(img_list)):
        curr_id = imgs['C'][x+1].value
        curr_img_url_list = []
        curr_img_url_list.append(str(imgs['B'][x+1].value))
        imgs['E'][x+1].value = None
        print(curr_id)
        for y in range(0, len(img_list)):
            if curr_id == imgs['C'][y+1].value and imgs['B'][y+1].value not in curr_img_url_list:
                curr_img_url_list.append(str(imgs['B'][y+1].value))
        for e in range (0, len(curr_img_url_list)):
            if e == 0:
                imgs['E'][x+1].value = str(curr_img_url_list[e])
            else:
                imgs['E'][x+1].value = str(str(imgs['E'][x+1].value) + ',' + str(curr_img_url_list[e]))
        print('url list done for: ' + str(imgs['C'][x+1].value))



def prod_img_urls():
    for x in range(1, len(imgs['A'])):
        oldImg = str(imgs['A'][x].value)
        imgs['B'][x].value = str('https://ukens-dental.de/img/nordent_de_prod_images/original_images/' + oldImg)
            # print(str(imgs['B'][x].value) + '.....' + str(imgs['A'][x].value))


def run_img_funs():
    #imgs_to_prods()
    #prod_img_urls()
    #all_lists_and_names()
    #all_imgs_urls_to_single_product_id()
    single_urls()
    wb_imgs.save('files\products_images.xlsx')

wb_prod_imgs_2 = openpyxl.load_workbook('files\products_images.xlsx')
url_lists = wb_prod_imgs_2['Sheet1']

wb_prod_img_urls = openpyxl.load_workbook('files\single_prod_img_urls.xlsx')
prod_img_urls = wb_prod_img_urls['Sheet1']


def single_urls():
    
    id_list = []

    print('running single url function...')

    for x in range (0, len(prod_img_urls['A'])):
        if url_lists['C'][x+1].value not in id_list:
            prod_img_urls['A'][x+1].value = url_lists['C'][x+1].value
            prod_img_urls['B'][x+1].value = url_lists['E'][x+1].value
            id_list.append(url_lists['A'][x+1].value)
        wb_prod_img_urls.save('files\single_prod_img_urls.xlsx')

    print('single url function done.')
    print(id_list)


#run_img_funs()
print('image functions done.')

# End image functions

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

# Begin product functions

print('')

print('running product functions on products.xlsx ...')  

wb_prods = openpyxl.load_workbook('files\products.xlsx')
prods = wb_prods['Sheet1']

# Rewrite Categories to be written out by name, no commas in name
def prods_corr_cat_ids():
    # Load in categories, and makes parallel lists of the ids and names for checking
    cats_f_wb = openpyxl.load_workbook('files\cats_edited.xlsx')
    cats_for_prods = cats_f_wb['Sheet1']

    cat_id = [] 
    cat_names = []

    for a in range(1, len(cats_for_prods['A'])):
        cat_id.append(cats_for_prods['A'][a].value)
        cat_names.append(cats_for_prods['G'][a].value)
    print(cat_id)
    print(cat_id[0])
    print(len(cat_id))
    print(cat_names)
    print(cat_names[0])
    print(len(cat_names))

    # increase id by 1000, so same as cats list
    for x in range(0, len(prods['B'])-1):
        #print(str(prods['B'][x+1].value))
        prods['B'][x+1].value = int(int(prods['B'][x+1].value +1000))
        # for each value in column b (x from larger loop used), if same as some value y in
        # cat_id list, change number to name in prods wb, and remove commas which interfere
        # with csv uploading
        for y in range(0, len(cat_id)):
            if prods['B'][x+1].value == cat_id[y]:
                prods['B'][x+1].value = cat_names[y]
                if ',' in str(prods['B'][x+1].value):
                    prods['B'][x+1].value = str(prods['B'][x+1].value).replace(',', '')
                #print(str(prods['B'][x+1].value))
        # save every time as code ends in an error
        #wb_prods.save('files\products_post.xlsx')

    #print(str(prods['B'][x+1].value))

    
    print('increasing prods cat ids by 1000 done')
    wb_prods.save('files\products.xlsx')
    print('cat chanes now saved to products.xlsx')


# must remove # and ; from name column E (pound and semicolon symbols) from product names, for PS NOT NECESSARY, AS REPLACED IN 'products.xlsx' FILE
def prods_corr_symbols():
    for x in range(1, len(prods['E'])):
        #print(str(prods['E'][x].value))
        if '#' in str(prods['E'][x].value):
            print('this title with # is up: ' + str(prods['E'][x].value))
            prods['E'][x].value = str(prods['E'][x].value).replace(',', 'Nr. ')
            print('the title is now: ' + str(prods['E'][x].value))
        if ';' in str(prods['E'][x].value):
            print('this title with # is up: ' + str(prods['E'][x].value))
            prods['E'][x].value = str(prods['E'][x].value).replace(';', '')
            print('the title is now: ' + str(prods['E'][x].value))

# alt tags for images and product tags
def prods_tags():
    prods['J'][0].value = str('image_alt_tags')
    prods['K'][0].value = str('prod_alt_tags')
    for x in range(1, len(prods['E'])):
        # string of category and product name with some filler for alt image tags
        prods_tags = str(str(prods['E'][x].value) 
        + ' von Nordent ist in Kategorie ' + 
        str(prods['B'][x].value) + 
        ' und wird angeboten von Ukens Dental')
        # if commas in size of something, replace with . to maintain size
        pattern = re.compile(r'(?<=\d),(?=\d)')
        pattern.sub('.',prods_tags)
        # otherwise, delete all commas as each images' tags are separated by so
        if ',' in prods_tags:
            prods_tags = prods_tags.replace(',', '')
        prods['J'][x].value = prods_tags
        #print(prods['J'][x].value)
        # determine how many images per product need an alt description, make as many as needed by count of commas, now up to 3
        com_count = str(prods['I'][x].value)
        com_count.count(',')
        if com_count.count(',') == 1:
            prods['J'][x].value = str(prods['J'][x].value + ',' + prods['J'][x].value)
        if com_count.count(',') == 2:
            prods['J'][x].value = str(prods['J'][x].value + ',' + prods['J'][x].value + ',' + prods['J'][x].value)
        # now the actual product tags, which is category, product, nordent and ukens dental, Dentalinstrumente
        prod_alt_tags = str(str(prods['E'][x].value) 
        + ',Nordent,' + 
        str(prods['B'][x].value) + 
        ',Ukens Dental' + ',Dentalinstrumente')
        pattern = re.compile(r'(?<=\d),(?=\d)')
        pattern.sub('.',prod_alt_tags)
        prods['K'][x].value = prod_alt_tags
        print(prods['K'][x].value)



# OUT OF CONVENIENCE SAVING CHANGES TO products.xlsx !!!! so don't have to run corr_cat_ids
def run_prod_funs():
    #prods_corr_cat_ids()
    #prods_corr_symbols()
    prods_tags()
    wb_prods.save('files\products.xlsx')

run_prod_funs()
print('product functions done.')

